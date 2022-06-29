import openpyxl
import orjson as json
import requests
from loguru import logger

PAT = 'v9PYQsqTgtxyEiidQezH'
PROJECT_ID = 273

url = f"https://git.minerva.vn/api/v4/projects/{PROJECT_ID}/issues"

headers = {
    'PRIVATE-TOKEN': PAT,
    'Content-Type': 'application/json'
}


def generate_issue(data):
    res = requests.request("POST", url, headers=headers, data=json.dumps(data))
    logger.info(res.text)


def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value


def generate_issue_from_excel():
    wb = openpyxl.load_workbook(filename='./APIS.xlsx')
    ws = wb['FUNCTION LIST']
    for index, row in enumerate(ws.iter_rows()):
        if 5 < index < 110 and row[0].value is not None:
            data = {
                'title': f"[{getMergedCellVal(ws, row[1])}] {row[2].value}",
                'description': f"- WS/Function: {row[3].value}\n - DB: {row[4].value}\n - ServiceDomain: {row[5].value}\n - Tên hàm: {row[6].value}\n - REST/JSON: {row[7].value}\n - DESCRIPTION: {row[8].value}",
                "milestone_id": 103
            }
            logger.info(data['title'], index)
            generate_issue(data=data)


if __name__ == '__main__':
    generate_issue_from_excel()
